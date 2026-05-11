"""Export utilities for DataVaidya.

Provides exporters for CSV, Excel, Parquet, ydata-profiling HTML,
PDF executive summaries, and a self-contained Python replay script
generated from a cleaning log.
"""
from __future__ import annotations

import io
import logging
import textwrap
from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd

from utils.constants import APP_NAME, APP_VERSION

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def export_csv(df: pd.DataFrame) -> BytesIO:
    """Export a DataFrame to CSV (UTF-8 with BOM, comma-separated).

    Returns a BytesIO positioned at the start. Empty DataFrames still
    produce a valid CSV containing the header row only.
    """
    buf = BytesIO()
    # utf-8-sig writes a BOM, helping Excel detect UTF-8 correctly.
    text = df.to_csv(index=False, sep=",")
    buf.write(text.encode("utf-8-sig"))
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def export_excel(df: pd.DataFrame, sheet_name: str = "cleaned") -> BytesIO:
    """Export a DataFrame to an .xlsx workbook using openpyxl.

    Column widths are auto-fit to ``min(50, max(header_len, sample_max_len))``
    where the sample is the first 100 non-null values per column.
    Sheet names longer than the Excel 31-character limit are truncated.
    """
    safe_sheet = (sheet_name or "Sheet1")[:31] or "Sheet1"
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=safe_sheet, index=False)
        worksheet = writer.sheets[safe_sheet]
        for idx, col in enumerate(df.columns, start=1):
            header_len = len(str(col))
            try:
                sample = df[col].dropna().head(100).astype(str)
                sample_max_len = int(sample.map(len).max()) if not sample.empty else 0
            except (TypeError, ValueError, AttributeError):
                sample_max_len = 0
            width = min(50, max(header_len, sample_max_len) + 2)
            # openpyxl column letters: 1 -> 'A', 27 -> 'AA', etc.
            from openpyxl.utils import get_column_letter
            worksheet.column_dimensions[get_column_letter(idx)].width = width
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Parquet
# ---------------------------------------------------------------------------
def export_parquet(df: pd.DataFrame) -> BytesIO:
    """Export a DataFrame to Parquet with Snappy compression.

    If the write fails because of unhashable / non-serializable object
    columns, those object columns are cast to ``str`` and the write is
    retried (a warning is logged).
    """
    buf = BytesIO()
    try:
        df.to_parquet(buf, index=False, compression="snappy")
    except (TypeError, ValueError, pd.errors.PerformanceWarning) as exc:
        log.warning(
            "Parquet write failed (%s); casting object columns to str and retrying.",
            exc,
        )
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].astype(str)
        buf = BytesIO()
        df.to_parquet(buf, index=False, compression="snappy")
    except Exception as exc:  # noqa: BLE001 - last-resort retry
        log.warning(
            "Unexpected parquet write failure (%s); casting object columns to str.",
            exc,
        )
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].astype(str)
        buf = BytesIO()
        df.to_parquet(buf, index=False, compression="snappy")
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# ydata-profiling HTML
# ---------------------------------------------------------------------------
def _profile_cache_key(df: pd.DataFrame, title: str, minimal: bool) -> tuple[Any, ...]:
    """Build a stable cache key for the profile report based on data identity."""
    return (df.shape, tuple(df.columns), df.head(50).to_csv(), title, minimal)


def _render_profile_html(
    df: pd.DataFrame, title: str = "DataVaidya Profile", minimal: bool = True
) -> BytesIO:
    """Inner implementation of the profile report exporter.

    Separated from the Streamlit-decorated entry point so that the cache
    decorator can be applied lazily.
    """
    try:
        from ydata_profiling import ProfileReport
    except ImportError as exc:
        log.warning("ydata_profiling unavailable: %s", exc)
        html = (
            "<!doctype html><html><head><meta charset='utf-8'>"
            f"<title>{title}</title></head><body>"
            "<h1>Profile report unavailable</h1>"
            f"<p>The optional dependency <code>ydata-profiling</code> is not "
            f"installed: {exc}.</p>"
            "</body></html>"
        )
        buf = BytesIO(html.encode("utf-8"))
        buf.seek(0)
        return buf

    target = df
    try:
        report = ProfileReport(target, title=title, minimal=minimal, progress_bar=False)
        html = report.to_html()
    except MemoryError:
        log.warning(
            "MemoryError while profiling full dataset; falling back to a 10k-row sample."
        )
        try:
            sample = df.sample(min(10_000, len(df)), random_state=42)
        except ValueError:
            sample = df
        report = ProfileReport(
            sample,
            title=f"{title} (sampled 10k rows)",
            minimal=True,
            progress_bar=False,
        )
        html = report.to_html()
    except Exception as exc:  # noqa: BLE001 - convert to inline error page
        log.warning("Profile generation failed: %s", exc)
        html = (
            "<!doctype html><html><head><meta charset='utf-8'>"
            f"<title>{title}</title></head><body>"
            "<h1>Profile report failed</h1>"
            f"<pre>{exc}</pre></body></html>"
        )

    buf = BytesIO(html.encode("utf-8"))
    buf.seek(0)
    return buf


def export_profile_html(
    df: pd.DataFrame, title: str = "DataVaidya Profile", minimal: bool = True
) -> BytesIO:
    """Generate a ydata-profiling HTML report as a BytesIO.

    Wrapped with ``st.cache_data`` (when Streamlit is available), keyed
    on ``(df.shape, tuple(df.columns), df.head(50).to_csv())`` so that
    re-running on identical input data is a cache hit.

    On MemoryError, falls back to ``df.sample(10_000, random_state=42)``
    and logs a warning. ``minimal=True`` is the default to keep memory
    use manageable on a 1 GB RAM tier.
    """
    try:
        import streamlit as st

        @st.cache_data(show_spinner=False)
        def _cached(
            shape: tuple[int, int],
            columns: tuple[str, ...],
            head_csv: str,
            title_in: str,
            minimal_in: bool,
        ) -> bytes:
            # Closure over df: cache key is the explicit args, but the
            # actual data used is the live df (the args uniquely identify
            # the head/shape/columns we care about).
            del shape, columns, head_csv  # only used for cache identity
            return _render_profile_html(df, title=title_in, minimal=minimal_in).getvalue()

        key = _profile_cache_key(df, title, minimal)
        payload = _cached(key[0], key[1], key[2], title, minimal)
        buf = BytesIO(payload)
        buf.seek(0)
        return buf
    except ImportError:
        # Streamlit not installed: just call through directly.
        return _render_profile_html(df, title=title, minimal=minimal)


# ---------------------------------------------------------------------------
# PDF summary
# ---------------------------------------------------------------------------
def export_pdf_summary(
    markdown_text: str, title: str = "DataVaidya Executive Summary"
) -> BytesIO:
    """Render a markdown string to a PDF via fpdf2 + markdown-it-py.

    A4 page, 15 mm margins, Helvetica. Adds a title block and a
    "Generated YYYY-MM-DD" line at the top. Characters that Helvetica
    cannot encode (e.g. astral-plane unicode) are replaced with ``'?'``
    and a warning is logged.
    """
    from fpdf import FPDF
    from markdown_it import MarkdownIt

    md = MarkdownIt("commonmark", {"html": True}).enable("table")
    body_html = md.render(markdown_text or "")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=18)

    generated_at = datetime.now().strftime("%Y-%m-%d")

    def _safe(text: str) -> str:
        """Coerce text into the Latin-1 subset Helvetica supports."""
        try:
            text.encode("latin-1")
            return text
        except UnicodeEncodeError:
            log.warning(
                "PDF: replacing unencodable characters in text of length %d.",
                len(text),
            )
            return text.encode("latin-1", errors="replace").decode("latin-1")

    try:
        pdf.cell(0, 10, _safe(title), new_x="LMARGIN", new_y="NEXT")
    except TypeError:
        # Older fpdf2 signatures fall back to ln=
        pdf.cell(0, 10, _safe(title), ln=1)

    pdf.set_font("Helvetica", size=10)
    try:
        pdf.cell(0, 6, _safe(f"Generated {generated_at}"), new_x="LMARGIN", new_y="NEXT")
    except TypeError:
        pdf.cell(0, 6, _safe(f"Generated {generated_at}"), ln=1)
    pdf.ln(4)

    pdf.set_font("Helvetica", size=11)
    safe_html = _safe(body_html)
    try:
        pdf.write_html(safe_html)
    except Exception as exc:  # noqa: BLE001 - fall back to plain-text dump
        log.warning("pdf.write_html failed (%s); falling back to multi_cell.", exc)
        pdf.multi_cell(0, 6, _safe(markdown_text or ""))

    out = pdf.output(dest="S")
    # fpdf2 returns bytearray; older fpdf returns str.
    if isinstance(out, str):
        payload = out.encode("latin-1", errors="replace")
    else:
        payload = bytes(out)
    buf = BytesIO(payload)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Python script generator
# ---------------------------------------------------------------------------
OP_DEFS: dict[str, str] = {
    "fill_missing": '''
def op_fill_missing(df, columns=None, strategy="median", value=None):
    cols = columns if columns is not None else df.columns.tolist()
    for c in cols:
        if c not in df.columns:
            continue
        if strategy == "mean":
            df[c] = df[c].fillna(df[c].mean())
        elif strategy == "median":
            df[c] = df[c].fillna(df[c].median())
        elif strategy == "mode":
            if not df[c].mode().empty:
                df[c] = df[c].fillna(df[c].mode().iloc[0])
        elif strategy == "constant":
            df[c] = df[c].fillna(value)
        elif strategy == "ffill":
            df[c] = df[c].ffill()
        elif strategy == "bfill":
            df[c] = df[c].bfill()
    return df
''',
    "drop_missing": '''
def op_drop_missing(df, columns=None, how="any", thresh=None):
    subset = columns if columns else None
    if thresh is not None:
        return df.dropna(subset=subset, thresh=thresh).reset_index(drop=True)
    return df.dropna(subset=subset, how=how).reset_index(drop=True)
''',
    "drop_duplicates": '''
def op_drop_duplicates(df, columns=None, keep="first"):
    subset = columns if columns else None
    return df.drop_duplicates(subset=subset, keep=keep).reset_index(drop=True)
''',
    "cap_outliers_iqr": '''
def op_cap_outliers_iqr(df, columns=None, k=1.5):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    for c in cols:
        if c not in df.columns:
            continue
        q1 = df[c].quantile(0.25)
        q3 = df[c].quantile(0.75)
        iqr = q3 - q1
        lower = q1 - k * iqr
        upper = q3 + k * iqr
        df[c] = df[c].clip(lower=lower, upper=upper)
    return df
''',
    "remove_outliers_iqr": '''
def op_remove_outliers_iqr(df, columns=None, k=1.5):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    mask = pd.Series(True, index=df.index)
    for c in cols:
        if c not in df.columns:
            continue
        q1 = df[c].quantile(0.25)
        q3 = df[c].quantile(0.75)
        iqr = q3 - q1
        lower = q1 - k * iqr
        upper = q3 + k * iqr
        mask &= df[c].between(lower, upper) | df[c].isna()
    return df.loc[mask].reset_index(drop=True)
''',
    "strip_whitespace": '''
def op_strip_whitespace(df, columns=None):
    cols = columns if columns is not None else df.select_dtypes(include="object").columns.tolist()
    for c in cols:
        if c in df.columns and df[c].dtype == object:
            df[c] = df[c].astype(str).str.strip()
    return df
''',
    "standardize_case": '''
def op_standardize_case(df, columns=None, case="lower"):
    cols = columns if columns is not None else df.select_dtypes(include="object").columns.tolist()
    for c in cols:
        if c not in df.columns or df[c].dtype != object:
            continue
        s = df[c].astype(str)
        if case == "lower":
            df[c] = s.str.lower()
        elif case == "upper":
            df[c] = s.str.upper()
        elif case == "title":
            df[c] = s.str.title()
        elif case == "capitalize":
            df[c] = s.str.capitalize()
    return df
''',
    "rename_columns": '''
def op_rename_columns(df, mapping=None):
    if mapping:
        df = df.rename(columns=dict(mapping))
    return df
''',
    "drop_columns": '''
def op_drop_columns(df, columns=None):
    if not columns:
        return df
    to_drop = [c for c in columns if c in df.columns]
    return df.drop(columns=to_drop)
''',
    "parse_dates": '''
def op_parse_dates(df, columns=None, format=None, errors="coerce"):
    cols = columns if columns is not None else []
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], format=format, errors=errors)
    return df
''',
    "cast_dtype": '''
def op_cast_dtype(df, columns=None, dtype="float"):
    cols = columns if columns is not None else []
    for c in cols:
        if c in df.columns:
            try:
                df[c] = df[c].astype(dtype)
            except (ValueError, TypeError):
                df[c] = pd.to_numeric(df[c], errors="coerce") if dtype in ("int", "float", "int64", "float64") else df[c]
    return df
''',
    "downcast_numeric": '''
def op_downcast_numeric(df, columns=None):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    for c in cols:
        if c not in df.columns:
            continue
        if pd.api.types.is_integer_dtype(df[c]):
            df[c] = pd.to_numeric(df[c], downcast="integer")
        elif pd.api.types.is_float_dtype(df[c]):
            df[c] = pd.to_numeric(df[c], downcast="float")
    return df
''',
    "clip_range": '''
def op_clip_range(df, columns=None, lower=None, upper=None):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    for c in cols:
        if c in df.columns:
            df[c] = df[c].clip(lower=lower, upper=upper)
    return df
''',
    "normalize_minmax": '''
def op_normalize_minmax(df, columns=None):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    for c in cols:
        if c not in df.columns:
            continue
        mn = df[c].min()
        mx = df[c].max()
        rng = mx - mn
        if rng and not pd.isna(rng):
            df[c] = (df[c] - mn) / rng
    return df
''',
    "standardize_zscore": '''
def op_standardize_zscore(df, columns=None):
    cols = columns if columns is not None else df.select_dtypes(include="number").columns.tolist()
    for c in cols:
        if c not in df.columns:
            continue
        mu = df[c].mean()
        sd = df[c].std()
        if sd and not pd.isna(sd):
            df[c] = (df[c] - mu) / sd
    return df
''',
    "filter_rows": '''
def op_filter_rows(df, query=None):
    if not query:
        return df
    try:
        return df.query(query).reset_index(drop=True)
    except (ValueError, SyntaxError, pd.errors.UndefinedVariableError) as exc:
        log.warning("filter_rows query failed (%s); returning df unchanged.", exc)
        return df
''',
    "reset_index_op": '''
def op_reset_index_op(df, drop=True):
    return df.reset_index(drop=drop)
''',
}


# Function name dispatched from cleaning log op id -> generated function.
_OP_TO_FUNC: dict[str, str] = {name: f"op_{name}" for name in OP_DEFS}


SCRIPT_TEMPLATE = '''"""
{header}
Generated: {generated_at}
Source app: {app_name} v{app_version}
Steps: {n_steps}

Usage: python clean.py <input> [--output cleaned.csv]
"""
from __future__ import annotations
import argparse
import logging
from pathlib import Path
import pandas as pd
import numpy as np

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("datavaidya")


def load(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".csv":   return pd.read_csv(path)
    if ext == ".tsv":   return pd.read_csv(path, sep="\\t")
    if ext in (".xlsx", ".xls"): return pd.read_excel(path)
    if ext == ".parquet": return pd.read_parquet(path)
    if ext == ".json":  return pd.read_json(path)
    raise ValueError(f"Unsupported extension: {{ext}}")


def save(df: pd.DataFrame, path: Path) -> None:
    ext = path.suffix.lower()
    if ext == ".csv":   df.to_csv(path, index=False)
    elif ext == ".parquet": df.to_parquet(path, index=False)
    elif ext in (".xlsx", ".xls"): df.to_excel(path, index=False)
    else: df.to_csv(path, index=False)


{op_definitions}


def run(df: pd.DataFrame) -> pd.DataFrame:
    log.info("Starting with %d rows x %d cols", *df.shape)
{run_body}
    log.info("Finished with %d rows x %d cols", *df.shape)
    return df


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="DataVaidya cleaning pipeline")
    parser.add_argument("input", help="Input file path")
    parser.add_argument("--output", default="cleaned.csv", help="Output path")
    args = parser.parse_args()
    df_in = load(Path(args.input))
    df_out = run(df_in)
    save(df_out, Path(args.output))
    log.info("Wrote %s", args.output)
'''


def _validate_column_name(name: Any) -> None:
    """Reject column names containing newline, carriage return or backtick."""
    if not isinstance(name, str):
        return
    for bad in ("\n", "\r", "`"):
        if bad in name:
            raise ValueError(
                f"Unsafe character {bad!r} in column name {name!r}; "
                "refusing to embed in generated script."
            )


def _validate_params(params: dict[str, Any]) -> None:
    """Recursively scan params for unsafe column-name-like strings."""
    for k, v in params.items():
        _validate_column_name(k)
        if isinstance(v, str):
            _validate_column_name(v)
        elif isinstance(v, (list, tuple, set)):
            for item in v:
                if isinstance(item, str):
                    _validate_column_name(item)
        elif isinstance(v, dict):
            for kk, vv in v.items():
                if isinstance(kk, str):
                    _validate_column_name(kk)
                if isinstance(vv, str):
                    _validate_column_name(vv)


def export_python_script(
    cleaning_log: list[tuple[str, dict]],
    source_filename: str = "input.csv",
    app_version: str = APP_VERSION,
) -> BytesIO:
    """Generate a self-contained Python script replaying ``cleaning_log``.

    The output script only depends on pandas, numpy, argparse, logging
    and pathlib. Only the op function definitions that are actually used
    in the log are embedded. All parameter values are serialized with
    ``repr`` to prevent code injection. Column names containing newline,
    carriage return or backtick characters cause a ``ValueError``.

    An empty ``cleaning_log`` yields a valid script whose ``run`` is a
    no-op pass-through.
    """
    # Filter to known ops; warn about unknown ones.
    used_ops: list[str] = []
    seen: set[str] = set()
    valid_steps: list[tuple[str, dict]] = []
    for op_name, params in cleaning_log:
        if not isinstance(op_name, str):
            log.warning("Skipping non-string op entry: %r", op_name)
            continue
        if op_name not in OP_DEFS:
            log.warning("Skipping unknown op %r (not in OP_DEFS).", op_name)
            continue
        params = params or {}
        if not isinstance(params, dict):
            log.warning("Skipping op %s with non-dict params: %r", op_name, params)
            continue
        _validate_params(params)
        valid_steps.append((op_name, params))
        if op_name not in seen:
            used_ops.append(op_name)
            seen.add(op_name)

    if used_ops:
        op_definitions = "\n".join(textwrap.dedent(OP_DEFS[op]).strip() + "\n"
                                   for op in used_ops)
    else:
        op_definitions = "# (no cleaning ops recorded)\n"

    # Build run body
    if valid_steps:
        lines: list[str] = []
        for i, (op_name, params) in enumerate(valid_steps, start=1):
            func = _OP_TO_FUNC[op_name]
            kwargs = ", ".join(f"{k}={v!r}" for k, v in params.items())
            call = f"df = {func}(df)" if not kwargs else f"df = {func}(df, {kwargs})"
            lines.append(f'    log.info("Step {i}/{len(valid_steps)}: {op_name}")')
            lines.append(f"    {call}")
        run_body = "\n".join(lines)
    else:
        run_body = "    pass  # no cleaning steps recorded; pass-through"

    header = f"DataVaidya cleaning pipeline replay for: {source_filename}"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    script = SCRIPT_TEMPLATE.format(
        header=header,
        generated_at=generated_at,
        app_name=APP_NAME,
        app_version=app_version,
        n_steps=len(valid_steps),
        op_definitions=op_definitions,
        run_body=run_body,
    )

    buf = BytesIO(script.encode("utf-8"))
    buf.seek(0)
    return buf
